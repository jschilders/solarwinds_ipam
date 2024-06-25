from solarwinds_ipam import IPAM, IpNodeStatus, SubnetType
from openpyxl import load_workbook
import os
import dotenv

from openpyxl.worksheet.worksheet import Worksheet
from typing import Iterator

#
# This program will try to extract all the information of a Q-Park/Wayne Williams IP Worksheet (one per location)
#


def truncate_table(table: list[tuple], stop_if_empty: bool = True, start_value: str = "") -> Iterator[tuple]:
    #
    # Return the rows of the table after the first occurence of 'start value' in the first column  (if 'start value' is not empty)
    # stop returning rows if the value in first column is empty (when 'stop_if_empty' is true)
    #
    # table:list[tuple]     Table to process. A list of tuples, one tuple per row
    # stop_if_empty:bool    Stop returning rows if the first cell of the row is empty
    # start_value:str       Skip returning rows until this value in encountered in the first cell of a row
    #
    # Returns an iterator providing a tuple for each valid row
    #
    start_value_seen = False
    for row in table:
        if start_value and row[0] == start_value:
            start_value_seen = True
        if start_value and not start_value_seen:
            continue
        if stop_if_empty and row[0] is None:
            break
        yield row


def extract_dict(sheet: Worksheet, min_row: int = 2, max_row: int = None, min_col: int = 2, max_col: int = 3, stop_if_empty: bool = True, start_value: str = "") -> dict[str, str]:
    #
    # Extract a 2-column key/value table from a excel worksheet
    #
    # sheet:Worksheet       Excel worksheet of class openpyxl.Worksheet
    # min_row:int           First row of the sheet to process
    # max_row:int           Last row of the sheet to process
    # min_col:int           First column of sheet to process
    # max_col:int           Last column of the sheet to process (should be min_col+1)
    # stop_if_empty:bool    stop returning rows when the first cell of the row is empty
    # start_value:str       Start returning rows when 'start_value' is first found in the first cell of the row
    #
    # returns a dict containing the valid rows, first column is key, second column is value
    #
    table = sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
    if stop_if_empty or start_value:
        table = truncate_table(table, stop_if_empty, start_value)
    return dict(table)


def extract_table(sheet: Worksheet, min_row: int = 2, max_row: int = None, min_col: int = 2, max_col: int = None, stop_if_empty: bool = True, start_value: str = "") -> list[tuple]:  #
    #
    # Extract a mutli-column table as a list of tuples
    #
    # sheet:Worksheet       Excel worksheet of class openpyxl.Worksheet
    # min_row:int           First row of the sheet to process
    # max_row:int           Last row of the sheet to process
    # min_col:int           First column of sheet to process
    # max_col:int           Last column of the sheet to process
    # stop_if_empty:bool    stop returning rows when the first cell of the row is empty
    # start_value:str       Start returning rows when 'start_value' is first found in the first cell of the row
    #
    #
    # returns a list of tuples, one tuple per valid rows
    #
    table = sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
    if stop_if_empty or start_value:
        table = truncate_table(table, stop_if_empty, start_value)
    return table


def open_workbook(filename):
    # filename = 'UKSHRS - IP address schema.xlsx',
    # filename = 'BETOCE - subnets.xlsx',
    return load_workbook(filename=filename, read_only=True, data_only=True)


def process_summary(worksheet):
    #
    # Extract the header with key/value pairs
    #
    header = extract_dict(worksheet, min_row=2, max_row=5)
    pf_name = header["Parking Facility name"]
    pf_shortname = header["Parking Facility short name"]
    pf_city = header["City"]
    pf_uid = header["Parking Facility ID"]

    print(f"Processing subnets for site {pf_shortname} ({pf_name}/{pf_city})")

    uri = ipam.ipgroups.get_uri(FriendlyName=pf_shortname)
    if not uri:
        print(f"Group {pf_shortname} not found")
    else:
        new_comment = f"{pf_city} - {pf_name} ({pf_uid})"
        current_comment = ipam.ipgroups.get_comment(FriendlyName=pf_shortname)

        if new_comment != current_comment:
            updates = {"Comments": new_comment}
            print(f"{pf_shortname:<15} Updates: {updates}")
            ipam.ipsubnet.update(uri, **updates)

    #
    # Extract the "site subnet" (a single key/value pair)
    #
    site_supernet = extract_dict(worksheet, min_row=7, max_row=7)["Site Subnet - network address"]
    site_supernet_address, site_supernet_cidr = site_supernet.split("/")

    uri = ipam.ipsubnet.get_uri(Address=site_supernet_address, CIDR=site_supernet_cidr, GroupType=SubnetType.Supernet)
    if not uri:
        print(f"Group {pf_shortname} not found")
    else:
        new_comment = f"Site Subnet {pf_shortname}"
        current_comment = ipam.ipgroups.get_comment(FriendlyName=f"{site_supernet_address}/{site_supernet_cidr}")

        if new_comment != current_comment:
            updates = {"Comments": new_comment}
            print(f"{site_supernet_address}/{site_supernet_cidr} Updates: {updates}")
            ipam.ipsubnet.update(uri, **updates)

        networks = extract_table(worksheet, 11, max_col=8)
        for subnet_name, subnet_address, _, _, subnet_cidr, _, subnet_vlan in networks:
            subnet_cidr = subnet_cidr.lstrip("/")
            subnet_vlan = str(subnet_vlan)
            params: dict = {"Address": subnet_address, "CIDR": subnet_cidr}
            result: list = ipam._build_query("IPAM.Subnet", ["Comments", "VLAN"], params)
            if result:
                comment, vlan = result[0].get("Comments"), result[0].get("VLAN")
                updates = {}
                if subnet_name != comment:
                    updates["Comments"] = subnet_name
                if subnet_vlan != vlan:
                    updates["VLAN"] = subnet_vlan or ""
                if updates:
                    print(f"{subnet_address:<15} Updates: {updates}")
                    uri = ipam.ipsubnet.get_uri(**params)
                    ipam.ipsubnet.update(uri, **updates)
            else:
                print(subnet_name, subnet_address, subnet_cidr)
                parent_id = ipam.ipsubnet.get_id(Address=site_supernet_address, CIDR=site_supernet_cidr)
                print(parent_id)
                ipam.ipsubnet.create(Address=subnet_address, CIDR=subnet_cidr, ParentId=parent_id, FriendlyName=f"{subnet_address}/{subnet_cidr}", Comments=subnet_name or "", VLAN=subnet_vlan or "")


def process_tab(worksheet):  # noqa: C901
    print(f"Processing tab: {worksheet.title}")
    #
    # Extract the header with key/value pairs
    #
    header = extract_dict(worksheet, 2, 8)
    subnet_name = header["Name"]
    subnet_vlan = str(header["VLAN:"])
    subnet_address = header["Subnet:"]
    subnet_gateway = header["Default gateway"]
    subnet_mask = header["Subnet mask"]  # noqa: F841
    subnet_cidr = int(header["Subnet short form"].lstrip("/"))

    params: dict = {"Address": subnet_address, "CIDR": subnet_cidr}
    result: list = ipam._build_query("IPAM.Subnet", ["Comments", "VLAN"], params)
    if not result:
        print(f"Subnet {subnet_address}/{subnet_cidr} not found")
    else:
        comment, vlan = result[0].get("Comments"), result[0].get("VLAN")
        updates = {}
        if subnet_name != comment:
            updates["Comments"] = subnet_name
        if subnet_vlan != vlan:
            updates["VLAN"] = subnet_vlan
        if updates:
            print(f"{subnet_address:<15} Updates: {updates}")
            uri = ipam.ipsubnet.get_uri(**params)
            ipam.ipsubnet.update(uri, **updates)

        #
        # Set default gateway
        #
        uri = ipam.ipaddress.get_uri(IPAddress=subnet_gateway)
        if uri:
            ip_node = ipam.ipaddress.read(uri)
            assert subnet_gateway == ip_node["IPAddress"]
            updates = {}
            if ip_node["Comments"] != "Default Gateway":
                updates["Comments"] = "Default Gateway"
            if ip_node["Status"] != int(IpNodeStatus.Used) and ip_node["Status"] != int(IpNodeStatus.Reserved):
                updates["Status"] = int(IpNodeStatus.Reserved)
            if updates:
                print(f"{subnet_gateway:<15} Updates: {updates}")
                ipam.ipaddress.update(uri, **updates)
        else:
            print(f"Default Gateway {subnet_gateway} not found")

        #
        # Extract the address table, from row 11 until we
        # encounter an empty row or there are no more rows.
        #
        address_table = extract_table(worksheet, 11)
        for ip_address, name, switch, switchport, comment in address_table:
            if any([name, switch, switchport, comment]):
                uri = ipam.ipaddress.get_uri(IPAddress=ip_address)
                if not uri:
                    print(f"IP Address {ip_address} not found")
                else:
                    ip_node = ipam.ipaddress.read(uri)
                    assert ip_address == ip_node["IPAddress"]
                    updates = {}
                    if ip_node["Status"] != int(IpNodeStatus.Used):
                        if comment == "DHCP" or name == "DHCP":
                            if ip_node["Status"] != int(IpNodeStatus.Transient):
                                updates["Status"] = int(IpNodeStatus.Transient)
                        else:
                            if ip_node["Status"] != int(IpNodeStatus.Reserved):
                                updates["Status"] = int(IpNodeStatus.Reserved)

                    comment = comment or ""
                    if switch:
                        comment += f" // Connected to switch {switch}"
                    if switchport:
                        comment += f" on port {switchport}"
                    if comment != ip_node["Comments"]:
                        updates["Comments"] = comment

                    name = name or ""
                    if name != ip_node["DnsBackward"]:
                        updates["DnsBackward"] = name
                    if updates:
                        print(f"{ip_address:<15} Updates: {updates}")
                        ipam.ipaddress.update(uri, **updates)


def process_workbook(workbook):
    #
    # Iterate over all the worksheets (tabs/pages) in the workbook
    #
    for worksheet in workbook:
        if worksheet.title in ["Switches", "Calcs", "Translations"]:
            #
            # Skip the "calcs" and "translations" sheets, they contain no info we need
            #
            continue

        elif worksheet.title == "Summary":
            #
            # The "Summary" sheet has a different layout
            #
            process_summary(worksheet)

        else:
            #
            # Process the 'regular' pages
            #
            process_tab(worksheet)


##########################################################################################################


# Load extra environment variables from .env file
dotenv.load_dotenv()

# Get variables from environment
USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")
SERVER = os.getenv("SERVER")

#
# Open a connection to Solarwinds IPAM
#

ipam = IPAM(server=SERVER, username=USERNAME, password=PASSWORD, verify=False)

#
# Open a workbook (Excel spreadsheet)
#
workbook = open_workbook(
    # filename = 'UKSHRS - IP address schema.xlsx'
    filename="BETOCE - subnets.xlsx"
)

process_workbook(workbook)
